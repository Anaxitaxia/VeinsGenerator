import tkinter as tk
# from tkinter import ttk
from PIL import ImageGrab, Image, ImageTk
from tkinter import messagebox as mb
import tkinter.filedialog as fd
import openpyxl


class Struct(object):
    pass


imgs = Struct()


class MainForm(tk.Frame):
    def __init__(self, master=None, parent=None):
        super().__init__(master)
        self.parent = parent
        self.pack()
        self.center_window()
        self.add_widgets()
        self.master.title('Создание шаблона рисунка вен')
        self.master.resizable(False, False)
        self.focus_set()

    def determine_sizes(self):
        w = self.master.winfo_width()
        h = self.master.winfo_height()
        x = self.master.winfo_x()
        y = self.master.winfo_y()
        sizes = [w, h, x, y]
        return sizes

    def center_window(self, w=1000, h=600):
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) / 2
        y = (sh - h) / 2 - 45
        self.master.geometry('%dx%d+%d+%d' % (w, h, x, y))

    def add_widgets(self):
        def print_screen():
            self.update_idletasks()
            x_r = self.master.winfo_x()
            y_r = self.master.winfo_y()
            x_l = x_r + self.master.winfo_width()
            y_l = y_r + self.master.winfo_height()
            picture = ImageGrab.grab(
                (x_r + 10, y_r + 1, x_l, y_l + 45)
            )
            picture.save("screen.png", "PNG")
            mb.showinfo("", "Снимок успешно сделан")

        main_menu = tk.Menu(self.master)
        self.master.config(menu=main_menu)
        # mode_menu = tk.Menu(main_menu, tearoff=0)
        # mode_menu.add_command(label="добавить участок", command=self.show_reg_add)
        # mode_menu.add_command(label="выбрать участок")  # , command=self.show_reg_chs)
        # mode_menu.add_command(label="рисование")  # , command=self.show_drawing)
        main_menu.add_command(label='Загрузить изображение', command=self.place_main_img)
        main_menu.add_command(label='Добавить участок', command=self.show_reg_add)
        # main_menu.add_cascade(label='Режим работы', menu=mode_menu)
        # main_menu.add_command(label='Сохранить')
        main_menu.add_command(label='Скриншот', command=print_screen)
        # main_menu.add_command(label='Справка')
        main_menu.add_command(label='Выход', command=quit)

    def place_main_img(self, file='', flag=0):
        global filename
        if file == '':
            ftypes = [('Изображения', '*.png'), ('Изображения', '*.jpg'), ('Все файлы', '*')]
            dlg = fd.Open(self, filetypes=ftypes)
            filename = dlg.show()
        if filename != '':
            pil_im = Image.open(filename)
            main_canv = tk.Canvas(self.master, width=pil_im.size[0], height=pil_im.size[1])
            imgs.image = ImageTk.PhotoImage(pil_im)
            main_canv.place(x=pil_im.size[0] * 0.5, y=pil_im.size[1] * 0.5)
            main_canv.create_image(
                -pil_im.size[0] + pil_im.size[0] * 1.5,
                -pil_im.size[1] + pil_im.size[1] * 1.5,
                image=imgs.image
            )
            if not flag:
                self.center_window(pil_im.size[0] * 2, pil_im.size[1] * 2)

    def show_reg_add(self):
        self.update()
        sizes = self.determine_sizes()
        if filename:
            RegionAdding(sizes, master=tk.Toplevel(), parent=MainForm)
        else:
            mb.showwarning("Предупреждение", "Изображение не выбрано")

    '''def show_reg_chs(self):
        if filename:
            RegionChoose(master=tk.Toplevel(), parent=self)
        else:
            mb.showwarning("Предупреждение", "Изображение не выбрано")

    def show_drawing(self):
        if filename:
            DrawingVein(master=tk.Toplevel(), parent=self)
        else:
            mb.showwarning("Предупреждение", "Изображение не выбрано")'''


class RegionAdding(tk.Frame):
    def __init__(self, form_size, master=None, parent=None):
        super().__init__(master)
        form_size[1] = form_size[1] + 100
        self.parent = parent
        self.pack()
        self.master.geometry('%dx%d+%d+%d' % tuple(form_size))
        self.create_widgets(form_size)
        self.master.title('Добавление участка')
        self.master.resizable(False, False)
        self.focus_set()
        self.lines = []
        self.x_reg_coords = []
        self.y_reg_coords = []
        self.mode = 'usual'

    def change_cursor(self, _):
        if self.mode == 'draw':
            self.reg_add_canv.config(cursor="tcross")
        elif self.mode == 'usual':
            self.reg_add_canv.config(cursor="")

    def press_reg(self, _):
        if self.mode == 'draw':
            self.reg_add_canv.bind("<B1-Motion>", self.draw_reg)

    def draw_reg(self, evt):
        if self.mode == 'draw':
            self.lines.append(
                self.reg_add_canv.create_line(evt.x-1, evt.y-1, evt.x+1, evt.y+1, fill="skyblue", width=3)
            )
            self.x_reg_coords.append(evt.x)
            self.y_reg_coords.append(evt.y)

    def create_widgets(self, list_size):
        def border_drawing():
            self.mode = 'draw'

        def usual_mode():
            self.mode = 'usual'

        def erase_border():
            self.mode = 'usual'
            for line in self.lines:
                self.reg_add_canv.delete(line)
            self.lines = []
            self.x_reg_coords = []
            self.y_reg_coords = []

        def press_ok():
            try:
                wb = openpyxl.load_workbook("VeinsCoordsV05.xlsx")
                sheet = wb.active
            except FileNotFoundError:
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.cell(row=1, column=1, value='Название региона')
                sheet.cell(row=1, column=2, value='Минимальная координата x')
                sheet.cell(row=1, column=3, value='Минимальная координата y')
                sheet.cell(row=1, column=4, value='Максимальная координата x')
                sheet.cell(row=1, column=5, value='Максимальная координата y')
                sheet.cell(row=1, column=6, value='Ширина вены')
            index = 2
            while sheet.cell(row=index, column=1).value:
                index += 1
            sheet.cell(row=index, column=1, value=reg_name.get())
            max_reg_x = max(self.x_reg_coords)
            min_reg_x = min(self.x_reg_coords)
            max_reg_y = max(self.y_reg_coords)
            min_reg_y = min(self.y_reg_coords)
            sheet.cell(row=index, column=2, value=min_reg_x)
            sheet.cell(row=index, column=3, value=min_reg_y)
            sheet.cell(row=index, column=4, value=max_reg_x)
            sheet.cell(row=index, column=5, value=max_reg_y)
            selected_reg_ind = index
            # while sheet.cell(row=selected_reg_ind, column=index).value or\
            #         sheet.cell(row=selected_reg_ind, column=index + 2).value:
            #     index += 1
            # if index == 6:
            #     sheet.cell(row=selected_reg_ind, column=index, value=brush_width)
            #     index += 1
            index_x = 8
            index_y = 9
            for x in self.x_reg_coords:
                sheet.cell(row=selected_reg_ind, column=index_x, value=x)
                index_x += 2
            for y in self.y_reg_coords:
                sheet.cell(row=selected_reg_ind, column=index_y, value=y)
                index_y += 2
            wb.save("VeinsCoordsV05.xlsx")
            self.lines = []
            self.x_reg_coords = []
            self.y_reg_coords = []

        reg_add_menu = tk.Menu(self.master)
        self.master.config(menu=reg_add_menu)
        reg_add_menu.add_command(label="Выход", command=self.exit_reg_add)
        reg_name_lbl = tk.Label(self.master, text="Название:")
        reg_name_lbl.place(x=list_size[0] / 2 - list_size[0] / 4, y=10)
        reg_name = tk.StringVar()
        reg_name_entr = tk.Entry(self.master, width=50, textvariable=reg_name)
        reg_name_entr.place(x=list_size[0] / 2 - list_size[0] / 9, y=10)
        usual_btn = tk.Button(self.master, text="Обычный режим", width=round(list_size[0] / 30), command=usual_mode)
        border_btn = tk.Button(self.master,
                               text="Выделить границы",
                               width=round(list_size[0] / 30),
                               command=border_drawing
                               )
        erase_btn = tk.Button(self.master, text="Стереть границы", width=round(list_size[0] / 30), command=erase_border)
        usual_btn.place(x=list_size[0] / 10, y=list_size[1] / 5)
        border_btn.place(x=list_size[0] / 2.5, y=list_size[1] / 5)
        erase_btn.place(x=list_size[0] - list_size[0] / 3.5, y=list_size[1] / 5)
        pil_im_add = Image.open(filename)
        self.reg_add_canv = tk.Canvas(self.master, width=pil_im_add.size[0], height=pil_im_add.size[1])
        imgs.image = ImageTk.PhotoImage(pil_im_add)
        self.reg_add_canv.place(x=pil_im_add.size[0] * 0.5, y=pil_im_add.size[1] * 0.5 + 70)
        self.reg_add_canv.create_image(
            -pil_im_add.size[0] + pil_im_add.size[0] * 1.5,
            -pil_im_add.size[1] + pil_im_add.size[1] * 1.5,
            image=imgs.image
        )
        ok_btn = tk.Button(self.master, text="Ok", width=round(list_size[0] / 25), command=press_ok)
        ok_btn.place(x=list_size[0] / 2 - list_size[0] / 7, y=list_size[1] - list_size[1] / 5)
        self.reg_add_canv.bind("<Motion>", self.change_cursor)
        self.reg_add_canv.bind("<Button>", self.press_reg)

    def exit_reg_add(self):
        self.master.destroy()
        self.parent.place_main_img(self.master, filename, 1)


'''class RegionChoose(tk.Frame):
    def __init__(self, master=None, parent=None):
        super().__init__(master)
        self.parent = parent
        self.pack()
        self.master.geometry('400x100+500+200')
        self.create_widgets()
        self.master.title('Выбор участка')
        self.master.resizable(False, False)
        self.focus_set()

    def create_widgets(self):
        def press_select():
            global selected_reg_ind
            selected_reg_ind = cb_box.current() + 2
            self.master.destroy()

        reg_chs_menu = tk.Menu(self.master)
        self.master.config(menu=reg_chs_menu)
        reg_chs_menu.add_command(label="Выход", command=self.exit_reg_chs)
        reg_name_lbl = tk.Label(self.master, text="Название:")
        reg_name_lbl.place(x=50, y=0)
        wb = openpyxl.load_workbook("VeinsCoordsV05.xlsx")
        sheet = wb.active
        reg_names = []
        index = 2
        while sheet.cell(row=index, column=1).value:
            reg_names.append(sheet.cell(row=index, column=1).value)
            index += 1
        cb_box = ttk.Combobox(self.master, values=reg_names, exportselection=0)
        cb_box.pack(side=tk.TOP)
        ok_btn = tk.Button(self.master, text="Ok", width=50, command=press_select)
        ok_btn.place(x=20, y=50)

    def exit_reg_chs(self):
        self.master.destroy()'''


'''class DrawingVein(tk.Frame):
    def __init__(self, master=None, parent=None):
        super().__init__(master)
        self.parent = parent
        self.pack()
        self.update()
        wb = openpyxl.load_workbook('VeinsCoords.xlsx')
        sheet = wb.active
        w = sheet.cell(selected_reg_ind, 4).value - sheet.cell(selected_reg_ind, 2).value + 200
        h = sheet.cell(selected_reg_ind, 5).value - sheet.cell(selected_reg_ind, 3).value + 120
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        x = (sw - w) / 2
        y = (sh - h) / 2
        self.lines = []
        self.x_vein_coords = []
        self.y_vein_coords = []
        self.master.geometry('%dx%d+%d+%d' % (w, h, x, y))
        self.create_widgets([w, h, x, y])
        self.master.title('Рисование вены')
        self.master.resizable(False, False)
        self.focus_set()

    def create_widgets(self, list_size):
        def press_erase():
            for line in self.lines:
                self.drawing_canv.delete(line)
            self.lines = []
            self.x_vein_coords = []
            self.y_vein_coords = []

        def press_ok():
            wb = openpyxl.load_workbook("VeinsCoords.xlsx")
            sheet = wb.active
            index = 2
            while sheet.cell(row=index, column=1).value:
                index += 1
            max_reg_x = max(self.x_vein_coords)
            min_reg_x = min(self.x_vein_coords)
            max_reg_y = max(self.y_vein_coords)
            min_reg_y = min(self.y_vein_coords)
            sheet.cell(row=index, column=2, value=min_reg_x)
            sheet.cell(row=index, column=3, value=min_reg_y)
            sheet.cell(row=index, column=4, value=max_reg_x)
            sheet.cell(row=index, column=5, value=max_reg_y)
            index = 6
            while sheet.cell(row=selected_reg_ind, column=index).value or\
                    sheet.cell(row=selected_reg_ind, column=index + 2).value:
                index += 1
            if index == 6:
                sheet.cell(row=selected_reg_ind, column=index, value=brush_width)
                index += 1
            index_x = index + 2
            index_y = index_x + 1
            for x in self.x_vein_coords:
                sheet.cell(row=selected_reg_ind, column=index_x, value=x)
                index_x += 2
            for y in self.y_vein_coords:
                sheet.cell(row=selected_reg_ind, column=index_y, value=y)
                index_y += 2
            wb.save("VeinsCoordsV05.xlsx")

        drawing_menu = tk.Menu(self.master)
        self.master.config(menu=drawing_menu)
        drawing_menu.add_command(label="Выход", command=self.exit_drawing)
        width_lbl = tk.Label(self.master, text="Толщина кисти:")
        width_lbl.place(x=50, y=0)
        self.width_list = [1, 2, 3, 4, 5]
        self.comb_box = ttk.Combobox(self.master, values=self.width_list, exportselection=0)
        self.comb_box.current(0)
        global brush_width
        brush_width = 1
        self.comb_box.bind("<<ComboboxSelected>>", self.determine_brush_width)
        self.comb_box.pack(side=tk.TOP)
        wb = openpyxl.load_workbook('VeinsCoords.xlsx')
        sheet = wb.active
        pil_im = Image.open(filename)
        self.drawing_canv = tk.Canvas(self.master, width=pil_im.size[0], height=pil_im.size[1])
        imgs.image = ImageTk.PhotoImage(pil_im)
        self.drawing_canv.place(x=pil_im.size[0] * 0.5, y=pil_im.size[1] * 0.5 + 70)
        self.drawing_canv.create_image(
            -pil_im.size[0] + pil_im.size[0] * 1.5,
            -pil_im.size[1] + pil_im.size[1] * 1.5,
            image=imgs.image
        )
        self.drawing_canv.config(cursor="tcross")
        erase_btn = tk.Button(self.master, text="Стереть", width=round(list_size[0] / 25), command=press_erase)
        erase_btn.place(x=10, y=list_size[1] - list_size[1] / 5)
        ok_btn = tk.Button(self.master, text="Ok", width=round(list_size[0] / 25), command=press_ok)
        ok_btn.place(x=list_size[0] - list_size[0] / 2.5, y=list_size[1] - list_size[1] / 5)
        self.drawing_canv.bind("<B1-Motion>", self.draw_vein)

    def determine_brush_width(self, _):
        global brush_width
        brush_width = self.width_list[self.comb_box.current()]

    def draw_vein(self, evt):
        self.lines.append(
            self.drawing_canv.create_line(evt.x-1, evt.y-1, evt.x+1, evt.y+1, fill="skyblue", width=brush_width)
        )
        self.x_vein_coords.append(evt.x)
        self.y_vein_coords.append(evt.y)

    def exit_drawing(self):
        self.master.destroy()
        self.parent.place_main_img(filename, 1)'''


# запускает приложение
def main():
    global filename
    filename = ''
    root = tk.Tk()
    root["bg"] = "LightSteelBlue3"
    app = MainForm(master=root)
    root.mainloop()


if __name__ == '__main__':
    main()
